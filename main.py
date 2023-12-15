from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

from pydantic import BaseModel
import openpyxl

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class Guest(BaseModel):
        phone: int
        contactPhone: int
        firstName: str
        lastName: str
        surName: str
        dateOfBorn: str
        placeOfBorn: str
        registAddress: str
        typeDoc: str
        issuedBy: str
        dateOfIssue: str
        codeOffice: str
        numberDoc: str
        seriaDoc: str

def get_date(date):
     perem = date.split('-')

     return f'{perem[2]}.{perem[1]}.{perem[0]}'



@app.post("/api/addData")
def add_guest(guest: Guest):
    guest = guest.dict()

    wb = openpyxl.load_workbook('./data/data.xlsx')
    sheet = wb.active
    next_row = sheet.max_row + 1
    


    sheet.cell(row=next_row, column=1, value=guest['phone'])
    sheet.cell(row=next_row, column=2, value=guest['lastName'])
    sheet.cell(row=next_row, column=3, value=guest['firstName'])
    sheet.cell(row=next_row, column=4, value=guest['surName'])
    sheet.cell(row=next_row, column=5, value=get_date(guest['dateOfBorn']))
    sheet.cell(row=next_row, column=6, value=guest['placeOfBorn'])
    sheet.cell(row=next_row, column=7, value=guest['registAddress'])
    sheet.cell(row=next_row, column=8, value=guest['typeDoc'])
    sheet.cell(row=next_row, column=9, value=get_date(guest['dateOfIssue']))
    sheet.cell(row=next_row, column=10, value=guest['issuedBy'])
    sheet.cell(row=next_row, column=11, value=guest['seriaDoc'])
    sheet.cell(row=next_row, column=12, value=guest['numberDoc'])
    sheet.cell(row=next_row, column=13, value=guest['codeOffice'])
    sheet.cell(row=next_row, column=14, value=guest['contactPhone'])

    wb.save('./data/data.xlsx')

    print('успешное добавление элемента')

    return {"status": "Гость добавлен успешно"}

@app.get("/api/downloadData")
def download_data():
    return FileResponse('./data/data.xlsx', filename='data.xlsx', headers={"Content-Disposition": "attachment"})